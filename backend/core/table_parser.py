"""Format-neutral delimited-text and XLSX table import."""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from enum import Enum
from pathlib import Path
from typing import Any, Iterable

import chardet
import numpy as np
import pandas as pd
from openpyxl import load_workbook


class HeaderMode(str, Enum):
    AUTO = "auto"
    FIRST_ROW = "first_row"
    NONE = "none"
    ROW = "row"


@dataclass(frozen=True)
class ImportOptions:
    sheet: str | None = None
    header_mode: HeaderMode = HeaderMode.AUTO
    header_row: int | None = None  # one-based when supplied by a user


@dataclass(frozen=True)
class SheetInfo:
    name: str
    empty: bool


@dataclass
class ParseResult:
    dataframe: pd.DataFrame
    metadata: dict[str, Any] = field(default_factory=dict)
    header_row: int | None = 0
    delimiter: str | None = ","
    encoding: str | None = "utf-8"
    is_raser_format: bool = False
    sheet: str | None = None
    column_types: dict[str, str] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)


@dataclass
class TableInspection:
    format: str
    sheets: list[SheetInfo]
    suggested_sheet: str | None
    suggested_header_row: int | None
    delimiter: str | None
    encoding: str | None
    columns: list[str]
    column_types: dict[str, str]
    preview: list[list[Any]]
    warnings: list[str]
    is_raser_format: bool = False


def _detect_encoding(path: Path, sample_size: int = 65536) -> str:
    raw = path.read_bytes()[:sample_size]
    if raw[:3] == b"\xef\xbb\xbf":
        return "utf-8-sig"
    if raw[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return "utf-16"
    detected = (chardet.detect(raw).get("encoding") or "utf-8").lower()
    encoding = detected.replace("ascii", "utf-8")
    try:
        raw.decode(encoding)
    except (UnicodeDecodeError, LookupError):
        return "latin-1"
    return encoding


def _detect_delimiter(lines: list[str]) -> str:
    if lines and re.match(r"sep=.", lines[0]):
        return lines[0][4]
    sample_lines = [line for line in lines if line.strip() and not line.startswith("#")]
    sample = "\n".join(sample_lines[-20:])
    try:
        return csv.Sniffer().sniff(sample, delimiters=";,\t|").delimiter
    except csv.Error:
        for delimiter in (";", ",", "\t", "|"):
            counts = [line.count(delimiter) for line in sample_lines[-20:]]
            if counts and max(counts) > 0 and len(set(counts)) == 1:
                return delimiter
    return ","


def _is_number(value: Any) -> bool:
    if isinstance(value, bool) or value is None or value == "":
        return False
    if isinstance(value, (int, float, np.number)):
        return True
    try:
        float(str(value).strip().replace(",", "."))
        return True
    except ValueError:
        return False


def _table_width(rows: list[list[Any]]) -> int:
    widths = [len(row) for row in rows if any(value not in (None, "") for value in row)]
    if not widths:
        return 0
    return max(widths)


def _detect_header_row(rows: list[list[Any]]) -> int | None:
    width = _table_width(rows)
    if width == 0:
        return None
    for index, row in enumerate(rows[:40]):
        if len(row) != width:
            continue
        values = [value for value in row if value not in (None, "")]
        if not values:
            continue
        numeric = sum(_is_number(value) for value in values)
        unique = len({str(value).strip() for value in values}) == len(values)
        if numeric < len(values) * 0.5 and unique:
            return index
    return None


def _unique_columns(values: Iterable[Any], width: int) -> list[str]:
    columns: list[str] = []
    counts: dict[str, int] = {}
    source = list(values)
    for index in range(width):
        raw = source[index] if index < len(source) else None
        base = str(raw).strip() if raw is not None else ""
        if not base or base.lower().startswith("unnamed:"):
            base = f"Column {index + 1}"
        count = counts.get(base, 0) + 1
        counts[base] = count
        columns.append(base if count == 1 else f"{base} ({count})")
    return columns


def _rectangular(rows: list[list[Any]]) -> tuple[list[list[Any]], list[str]]:
    width = max((len(row) for row in rows), default=0)
    warnings: list[str] = []
    if any(len(row) != width for row in rows):
        warnings.append("Rows have different numbers of cells; missing cells were left empty.")
    return [row + [None] * (width - len(row)) for row in rows], warnings


def _coerce_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
    result = dataframe.copy()
    for column in result.columns:
        series = result[column]
        non_empty = series.dropna()
        non_empty = non_empty[non_empty.astype(str).str.strip() != ""]
        if non_empty.empty:
            continue
        numeric = pd.to_numeric(non_empty, errors="coerce")
        if numeric.notna().all():
            result[column] = pd.to_numeric(series, errors="coerce")
            continue
        if all(isinstance(value, (date, datetime, pd.Timestamp)) for value in non_empty):
            result[column] = pd.to_datetime(series, errors="coerce")
    return result


def _column_types(dataframe: pd.DataFrame) -> dict[str, str]:
    types: dict[str, str] = {}
    for column in dataframe.columns:
        series = dataframe[column]
        non_empty = series.dropna()
        non_empty = non_empty[non_empty.astype(str).str.strip() != ""]
        if non_empty.empty:
            kind = "empty"
        elif pd.api.types.is_datetime64_any_dtype(series):
            kind = "datetime"
        elif pd.api.types.is_numeric_dtype(series):
            kind = "numeric"
        else:
            value_kinds = {
                "numeric" if _is_number(value) else "datetime"
                if isinstance(value, (date, datetime, pd.Timestamp))
                else "text"
                for value in non_empty
            }
            kind = next(iter(value_kinds)) if len(value_kinds) == 1 else "mixed"
        types[str(column)] = kind
    return types


def _metadata(rows: list[list[Any]], header_index: int | None) -> dict[str, Any]:
    metadata: dict[str, Any] = {}
    if header_index is None:
        return metadata
    for row in rows[:header_index]:
        values = [value for value in row if value not in (None, "")]
        if len(values) == 1 and ":" in str(values[0]):
            key, value = str(values[0]).split(":", 1)
            metadata[key.strip()] = value.strip()
    return metadata


def _build_result(
    rows: list[list[Any]],
    options: ImportOptions,
    *,
    delimiter: str | None,
    encoding: str | None,
    sheet: str | None,
    is_raser: bool = False,
) -> ParseResult:
    if not rows or not any(any(value not in (None, "") for value in row) for row in rows):
        raise ValueError("No table cells could be read.")

    if options.header_mode == HeaderMode.NONE:
        header_index = None
    elif options.header_mode == HeaderMode.FIRST_ROW:
        header_index = 0
    elif options.header_mode == HeaderMode.ROW:
        if options.header_row is None or options.header_row < 1 or options.header_row > len(rows):
            raise ValueError("Header row is outside the table.")
        header_index = options.header_row - 1
    else:
        header_index = _detect_header_row(rows)

    data_rows = rows if header_index is None else rows[header_index + 1 :]
    data_rows, warnings = _rectangular(data_rows)
    width = max((len(row) for row in data_rows), default=_table_width(rows))
    if width == 0:
        raise ValueError("No data rows could be read.")

    header = [] if header_index is None else rows[header_index]
    columns = _unique_columns(header, width)
    dataframe = _coerce_columns(pd.DataFrame(data_rows, columns=columns))
    return ParseResult(
        dataframe=dataframe,
        metadata=_metadata(rows, header_index),
        header_row=header_index,
        delimiter=delimiter,
        encoding=encoding,
        is_raser_format=is_raser,
        sheet=sheet,
        column_types=_column_types(dataframe),
        warnings=warnings,
    )


def _text_rows(path: Path) -> tuple[list[list[Any]], str, str, bool]:
    encoding = _detect_encoding(path)
    lines = path.read_text(encoding=encoding, errors="replace").splitlines()
    delimiter = _detect_delimiter(lines)
    rows = [list(row) for row in csv.reader(lines, delimiter=delimiter)]
    is_raser = len(lines) > 1 and lines[0].startswith("sep=") and bool(
        re.match(r"Log\s+V\d+\.\d+", lines[1])
    )
    return rows, delimiter, encoding, is_raser


def _xlsx_sheets(path: Path) -> tuple[list[SheetInfo], str | None]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: list[SheetInfo] = []
    suggested: str | None = None
    try:
        for worksheet in workbook.worksheets:
            empty = not any(
                any(value not in (None, "") for value in row)
                for row in worksheet.iter_rows(values_only=True)
            )
            sheets.append(SheetInfo(worksheet.title, empty))
            if suggested is None and not empty:
                suggested = worksheet.title
    finally:
        workbook.close()
    return sheets, suggested


def _xlsx_rows(path: Path, sheet: str | None) -> tuple[list[list[Any]], str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        selected = sheet or next(
            (
                worksheet.title
                for worksheet in workbook.worksheets
                if any(
                    any(value not in (None, "") for value in row)
                    for row in worksheet.iter_rows(values_only=True)
                )
            ),
            None,
        )
        if selected is None:
            raise ValueError("Workbook contains no non-empty sheets.")
        if selected not in workbook.sheetnames:
            raise ValueError(f"Sheet does not exist: {selected}")
        rows = [list(row) for row in workbook[selected].iter_rows(values_only=True)]
        return rows, selected
    finally:
        workbook.close()


def parse_table(
    file_path: str | Path, options: ImportOptions | None = None
) -> ParseResult:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")
    options = options or ImportOptions()
    if path.suffix.lower() == ".xlsx":
        rows, sheet = _xlsx_rows(path, options.sheet)
        return _build_result(
            rows, options, delimiter=None, encoding=None, sheet=sheet
        )
    if path.suffix.lower() not in {".csv", ".tsv", ".txt"}:
        raise ValueError(f"Unsupported table format: {path.suffix or 'unknown'}")
    rows, delimiter, encoding, is_raser = _text_rows(path)
    return _build_result(
        rows,
        options,
        delimiter=delimiter,
        encoding=encoding,
        sheet=None,
        is_raser=is_raser,
    )


def _preview_values(dataframe: pd.DataFrame, count: int = 100) -> list[list[Any]]:
    preview = dataframe.head(count).astype(object).where(dataframe.head(count).notna(), None)
    return [
        [
            value.isoformat() if isinstance(value, (date, datetime, pd.Timestamp)) else value
            for value in row
        ]
        for row in preview.values.tolist()
    ]


def inspect_table(file_path: str | Path) -> TableInspection:
    path = Path(file_path)
    if path.suffix.lower() == ".xlsx":
        sheets, suggested = _xlsx_sheets(path)
        result = parse_table(path, ImportOptions(sheet=suggested))
        format_name = "xlsx"
    else:
        sheets, suggested = [], None
        result = parse_table(path)
        format_name = "delimited"
    return TableInspection(
        format=format_name,
        sheets=sheets,
        suggested_sheet=suggested,
        suggested_header_row=(
            result.header_row + 1 if result.header_row is not None else None
        ),
        delimiter=result.delimiter,
        encoding=result.encoding,
        columns=list(result.dataframe.columns),
        column_types=result.column_types,
        preview=_preview_values(result.dataframe),
        warnings=result.warnings,
        is_raser_format=result.is_raser_format,
    )
