from __future__ import annotations

from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app import app
from core.table_parser import HeaderMode, ImportOptions
from dataset import build_dataset


def test_dataset_retains_every_column_but_only_numeric_signals(tmp_path: Path) -> None:
    path = tmp_path / "mixed.csv"
    path.write_text("number,text,empty\n1,alpha,\n2,beta,\n", encoding="utf-8")

    dataset = build_dataset(str(path), path.name)

    assert list(dataset.table.columns) == ["number", "text", "empty"]
    assert dataset.order == ["number"]
    assert [column["type"] for column in dataset.columns] == [
        "numeric",
        "text",
        "empty",
    ]
    assert dataset.preview["columns"] == ["number", "text", "empty"]


def test_inspect_then_load_headerless_csv() -> None:
    client = TestClient(app)

    inspected = client.post(
        "/dataset/inspect",
        files={"file": ("headerless.csv", b"1,alpha\n2,beta\n", "text/csv")},
    )

    assert inspected.status_code == 200
    inspection = inspected.json()
    assert inspection["token"]
    assert inspection["format"] == "delimited"
    assert inspection["suggestedHeaderRow"] is None
    assert inspection["preview"]["rows"][0] == [1, "alpha"]

    loaded = client.post(
        "/dataset/load",
        data={"token": inspection["token"], "headerMode": "none"},
    )

    assert loaded.status_code == 200
    body = loaded.json()
    assert body["preview"]["columns"] == ["Column 1", "Column 2"]
    assert [signal["name"] for signal in body["signals"]] == ["Column 1"]
    assert body["dataset"]["columns"][1]["type"] == "text"


def test_xlsx_inspection_returns_sheet_choices() -> None:
    workbook = Workbook()
    workbook.active.title = "Empty"
    data = workbook.create_sheet("Data")
    data.append(["value"])
    data.append([3])
    content = BytesIO()
    workbook.save(content)

    response = TestClient(app).post(
        "/dataset/inspect",
        files={
            "file": (
                "workbook.xlsx",
                content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["sheets"] == [
        {"name": "Empty", "empty": True},
        {"name": "Data", "empty": False},
    ]
    assert body["suggestedSheet"] == "Data"


def test_preview_refresh_respects_sheet_and_header_options() -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "First"
    first.append(["value"])
    first.append([1])
    second = workbook.create_sheet("Second")
    second.append(["note"])
    second.append(["Time", "Reading"])
    second.append([0, 4.5])
    content = BytesIO()
    workbook.save(content)
    client = TestClient(app)
    inspected = client.post(
        "/dataset/inspect",
        files={"file": ("workbook.xlsx", content.getvalue())},
    ).json()

    response = client.post(
        "/dataset/preview",
        data={
            "token": inspected["token"],
            "sheet": "Second",
            "headerMode": "row",
            "headerRow": "2",
        },
    )

    assert response.status_code == 200
    assert response.json()["preview"] == {
        "columns": ["Time", "Reading"],
        "rows": [[0, 4.5]],
    }
